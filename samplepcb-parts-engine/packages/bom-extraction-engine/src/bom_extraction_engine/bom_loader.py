# -*- coding: utf-8 -*-
"""엑셀 로더 — 레거시(load_excel_pandas) 방식 차용 + 견고성 보강.

xlsx/xlsm: openpyxl read_only + data_only(수식 대신 계산값)
      실패 시 calamine 폴백 (openpyxl은 read_only여도 스타일시트를 파싱해서
      비표준 extLst 속성이 있는 파일에서 죽는다 — 레거시도 못 읽던 파일)
xls : pandas + xlrd, 실패 시 calamine 폴백
csv/tsv/bom : stdlib csv (인코딩 폴백 + 구분자 추정 + 가변 열 허용)
헤더 추정 없이 header=None 원본 그리드 그대로 DataFrame으로 만든다.

원본(header_probing_claude) 대비 웹 이식 보정: 업로드 허용 확장자인
xlsm/tsv가 xlrd 분기로 낙하해 실패하던 라우팅을 명시 분기로 고쳤다.
"""
import csv as _csv
import io
import warnings
import zipfile
from pathlib import Path
from typing import List

import pandas as pd

warnings.filterwarnings("ignore", category=UserWarning)

# openpyxl은 read_only여도 스타일시트를 통째로 파싱한다 — 서식 오염으로
# styles.xml이 수 MB로 부푼 파일(실측: BOM_MINI SERVO, 압축 해제 10.1MB)은
# 시트당 ~3초가 걸린다. 이런 파일만 스타일을 읽지 않는 calamine으로 우회.
_STYLES_BLOAT_BYTES = 2 * 1024 * 1024

# 서식만 남은 빈 열(엑셀이 16,384열까지 <c> 요소를 남기는 파일)은 시트 XML 자체가
# 수백 MB로 부푼다 — openpyxl 은 max_col 을 줘도 행마다 전 셀을 파싱하므로
# 사실상 끝나지 않는다(실측: 12,176행 × 16,384열, 시트 XML 279MB → 17분 CPU·1.7GB
# 후에도 미완). calamine 은 같은 파일을 1.8초에 (12176, 7) 로 읽는다.
_SHEET_BLOAT_BYTES = 50 * 1024 * 1024


def _needs_calamine(path: str) -> bool:
    """openpyxl 이 실용적으로 못 읽는 파일을 미리 가른다.

    styles.xml 비대(서식 오염)와 시트 XML 비대(빈 열 잔재)는 원인이 다르지만
    처방이 같다 — 스타일을 읽지 않고 사용 범위만 보는 calamine 으로 우회한다.
    """
    try:
        with zipfile.ZipFile(path) as z:
            if z.getinfo("xl/styles.xml").file_size > _STYLES_BLOAT_BYTES:
                return True
    except (KeyError, zipfile.BadZipFile, OSError):
        return False
    try:
        with zipfile.ZipFile(path) as z:
            return any(
                info.file_size > _SHEET_BLOAT_BYTES
                for info in z.infolist()
                if info.filename.startswith("xl/worksheets/")
                and info.filename.endswith(".xml")
            )
    except (zipfile.BadZipFile, OSError):
        return False


def _styles_bloated(path: str) -> bool:
    """구 이름 보존 — 호출부 호환용 별칭."""
    return _needs_calamine(path)


class _SharedRowWidths(list):
    """`attrs` 에 실어도 pandas 연산을 느리게 만들지 않는 행 폭 목록.

    pandas 는 거의 모든 연산에서 `__finalize__` 로 `attrs` 를 **deepcopy** 한다. 행 수만큼
    긴 리스트를 그대로 담으면 셀 접근 한 번이 리스트 전체 복사가 된다 — 실측: 12,000행
    CSV 에서 `iat` 18,246회가 deepcopy 5,480만 회를 유발해 **84초**(같은 크기 xlsx 는
    attrs 가 없어 2.3초). 값은 로드 시점에 확정돼 이후 변하지 않으므로 공유해도 안전하다.
    `list` 를 상속해 기존 소비자(`isinstance(widths, list)`)와 계약이 같다.
    """

    def __deepcopy__(self, _memo: dict) -> "_SharedRowWidths":
        return self

    def __copy__(self) -> "_SharedRowWidths":
        return self


def _load_csv(path: str) -> pd.DataFrame:
    """행마다 열 수가 다른(ragged) CSV도 크래시 없이 읽는다.

    pandas read_csv(sep=None, engine='python')는 일관된 열 수를 요구해
    가변 열 CSV에서 ParserError로 죽는다 — 실제 오픈소스/고객 BOM에 흔한
    형태라 stdlib csv.reader로 읽고 최대 폭으로 패딩한다.
    빈 문자열 셀은 pandas NaN 동작과 맞추기 위해 None으로 치환한다.
    """
    raw = Path(path).read_bytes()
    text = None
    for enc in ("utf-8-sig", "cp949", "utf-16", "latin1"):
        try:
            text = raw.decode(enc)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if text is None:
        raise ValueError(f"CSV 인코딩을 결정할 수 없습니다: {path}")
    delimiter = "\t" if path.lower().endswith(".tsv") else ","
    try:
        delimiter = _csv.Sniffer().sniff(text[:8192], delimiters=",\t;").delimiter
    except _csv.Error:
        pass
    rows = list(_csv.reader(io.StringIO(text), delimiter=delimiter))
    if not rows:
        return pd.DataFrame()
    source_row_widths = _SharedRowWidths(len(row) for row in rows)
    width = max(len(r) for r in rows)
    data = [[(c if c != "" else None) for c in r] + [None] * (width - len(r))
            for r in rows]
    frame = pd.DataFrame(data, dtype=object)
    # Padding deliberately makes ragged rows rectangular. Keep their original
    # widths so the workbook layer can distinguish a truly empty trailing
    # column from an unquoted delimiter inside a data cell after header roles
    # are known.
    frame.attrs["source_row_widths"] = source_row_widths
    frame.attrs["source_delimiter"] = delimiter
    return frame


def get_sheet_names(path: str) -> List[str]:
    ext = path.lower().rsplit(".", 1)[-1]
    if ext in ("csv", "tsv", "bom"):
        return [ext]
    if ext in ("xlsx", "xlsm"):
        if _styles_bloated(path):
            return pd.ExcelFile(path, engine="calamine").sheet_names
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=path, read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception:
            return pd.ExcelFile(path, engine="calamine").sheet_names
    if ext == "xls":
        try:
            return pd.ExcelFile(path, engine="xlrd").sheet_names
        except Exception:
            return pd.ExcelFile(path, engine="calamine").sheet_names
    raise ValueError(f"지원하지 않는 파일 형식: {ext}")


def load_sheet(path: str, sheet_idx: int = 0) -> pd.DataFrame:
    """시트 하나를 원본 그리드 그대로 로드한다."""
    ext = path.lower().rsplit(".", 1)[-1]
    if ext in ("csv", "tsv", "bom"):
        if sheet_idx != 0:
            raise ValueError("CSV/TSV/BOM은 시트가 하나입니다")
        return _load_csv(path)
    if ext in ("xlsx", "xlsm"):
        if _styles_bloated(path):
            return pd.read_excel(path, sheet_name=sheet_idx, header=None,
                                 engine="calamine")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            names = wb.sheetnames
            if sheet_idx >= len(names):
                raise ValueError(f"시트 인덱스 {sheet_idx} 초과 (총 {len(names)}개)")
            sheet = wb[names[sheet_idx]]
            data = [list(row) for row in sheet.iter_rows(values_only=True)]
            wb.close()
            return pd.DataFrame(data)
        except ValueError:
            raise
        except Exception:
            return pd.read_excel(path, sheet_name=sheet_idx, header=None,
                                 engine="calamine")
    if ext == "xls":
        try:
            return pd.read_excel(path, sheet_name=sheet_idx, header=None,
                                 engine="xlrd")
        except Exception:
            return pd.read_excel(path, sheet_name=sheet_idx, header=None,
                                 engine="calamine")
    raise ValueError(f"지원하지 않는 파일 형식: {ext}")

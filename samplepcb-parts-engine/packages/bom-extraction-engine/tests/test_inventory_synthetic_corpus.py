"""협력사 재고표 합성 코퍼스 — 엑셀·브로커 시트의 **알려진 병리**를 섞어 둔다.

왜 합성인가, 그리고 무엇을 못 하는가
------------------------------------
실물 협력사 파일은 아직 한 본(`EUREKA-stock parts 8.6.xlsx`)뿐이다. "한 파일로 설계한
추출기는 다음 파일에서 깨진다"는 게 코퍼스가 필요한 이유인데, 합성 파일은 결국 우리가
상상한 것만 담으므로 **깨질 지점을 발견하지는 못한다**. 회귀 방지용이다 — 여기 담긴 병리는
전부 실무에서 반복되는 것들이고, 한 번 고친 뒤 다시 깨지지 않게 못 박는 데 쓴다.
발견은 실물 두 번째 파일이 와야 한다(docs/PARTNER_PARTS.md §9).

병리를 파일마다 하나씩 두지 않고 **섞어 둔** 이유: 실제 파일이 그렇게 오고, 겹칠 때
나오는 상호작용이 진짜 사고 지점이다(예: 헤더가 5행 아래인 데다 품번이 날짜로 변한 시트).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from bom_extraction_engine import build_inventory_result


def _run(path: Path) -> dict:
    return build_inventory_result(
        input_path=path,
        original_filename=path.name,
        progress=lambda *_: None,
        config=None,
    )


def _rows(result: dict) -> list[dict]:
    return list(result.get("rows", []))


def _by_mpn(result: dict) -> dict[str, dict]:
    return {str(row.get("part_number") or ""): row for row in _rows(result)}


# ── 1호 — 서식 사고가 난 시트 ────────────────────────────────────────────────
# 엑셀이 파일을 **이미 망가뜨린 채** 도착하는 경우다. 원문 보존만으로는 복구되지 않으므로
# 최소한 "사람이 확인해야 한다"는 신호는 남아야 한다.
def _write_format_accidents(tmp_path: Path) -> Path:
    path = tmp_path / "broker-format-accidents.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "STOCK LIST"
    # 헤더가 1행이 아니다 — 로고·연락처가 위에 얹혀 있다(브로커 시트의 기본값에 가깝다).
    sheet.append(["ACME COMPONENTS CO., LTD"])
    sheet.append(["TEL 02-000-0000  /  sales@example.com"])
    sheet.append([])
    sheet.append(["Parts No.", "Brand", "QTY", "D/C"])
    # 엑셀이 날짜로 바꿔 버린 품번(`1-10` → 날짜). 문자열로 들어와도 표기가 이미 다르다.
    sheet.append(["Jan-10", "TI", 500, "23+"])
    # 앞자리 0 손실 — 숫자 서식이 먹었다.
    sheet.append([402, "Yageo", 10_000, "24+"])
    # 지수 표기로 굳은 긴 숫자 품번.
    sheet.append([1e5, "Samsung", 2_000, "24+"])
    # 정상 행 — 위 사고들이 정상 행까지 오염시키지 않는지 본다.
    sheet.append(["STM32F030F4P6", "ST", 1_200, "23+"])
    workbook.save(path)
    return path


def test_format_accidents_are_kept_and_flagged(tmp_path: Path):
    result = _run(_write_format_accidents(tmp_path))
    rows = _rows(result)

    # ① 헤더가 4행이어도 표를 찾는다 — 못 찾으면 시트가 통째로 사라진다.
    assert len(rows) == 4, f"헤더 탐지 실패로 행이 유실됐다: {rows}"

    # ② 정상 행은 사고 행 옆에서도 멀쩡하다.
    normal = _by_mpn(result).get("STM32F030F4P6")
    assert normal is not None
    assert normal.get("manufacturer") == "ST"
    assert normal.get("stock_qty") == 1_200

    # ③ 망가진 품번도 **버리지 않는다**. 복구는 못 해도 원문은 남고 재고는 살아 있다.
    #    (사람이 고칠 수 있게 하는 것이 이 프로필의 계약이다 — 행 수정 기능이 그 창구다.)
    quantities = sorted(row.get("stock_qty") for row in rows)
    assert quantities == [500, 1_200, 2_000, 10_000]
    for row in rows:
        assert (row.get("part_number_raw") or "") != "", f"원문이 비었다: {row}"


# ── 2호 — 한 시트에 표가 둘, 소계 행이 섞인 시트 ─────────────────────────────
def _write_two_tables(tmp_path: Path) -> Path:
    path = tmp_path / "broker-two-tables.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "재고"
    sheet.append(["품번", "메이커", "수량", "비고"])
    sheet.append(["MCP1700T-3302E/TT", "Microchip", "1,000", ""])
    sheet.append(["PIC16F1825T-I/SL", "Microchip", "2,500", ""])
    sheet.append(["소계", "", "3,500", ""])          # 합계 행 — 부품이 아니다
    sheet.append([])                                  # 구역 구분
    sheet.append(["[ 신규 입고 ]"])                   # 구역 제목
    sheet.append(["품번", "메이커", "수량", "비고"])  # 두 번째 표의 머리
    sheet.append(["CS5532-ASZR", "Cirrus", "700 pcs", "신규"])
    sheet.append(["LM337IMP", "TI", "1k", ""])
    workbook.save(path)
    return path


def test_second_table_is_not_lost_and_totals_are_visible(tmp_path: Path):
    result = _run(_write_two_tables(tmp_path))
    found = _by_mpn(result)

    # ① 두 번째 표가 통째로 유실되면 안 된다 — 브로커 시트는 구역을 나눠 쓴다.
    assert "CS5532-ASZR" in found, f"두 번째 표가 유실됐다: {sorted(found)}"
    assert "LM337IMP" in found

    # ② 첫 표도 그대로.
    assert found.get("MCP1700T-3302E/TT", {}).get("stock_qty") == 1_000

    # ③ 소계 행·시트 중간에 다시 나온 머리글은 부품이 아니다. 버리지는 않되(무유실)
    #    **검토 표시**가 있어야 관리자가 지울 수 있다 — 없으면 재고가 조용히 부풀고,
    #    원장 목록에 정체불명 행이 남는다(합성 코퍼스가 잡은 결함).
    for junk in ("소계", "품번", "[ 신규 입고 ]"):
        row = _by_mpn(result).get(junk)
        assert row is not None, f"{junk} 행이 사라졌다 — 무유실 위반"
        assert "mpn_needs_review" in (row.get("flags") or []), (
            f"{junk} 가 검토 표시 없이 부품으로 앉았다"
        )


def test_quantity_notations_are_read(tmp_path: Path):
    """`1,000` · `700 pcs` · `1k` — 브로커가 수량을 쓰는 세 가지 방식."""
    result = _run(_write_two_tables(tmp_path))
    found = _by_mpn(result)
    assert found.get("MCP1700T-3302E/TT", {}).get("stock_qty") == 1_000
    assert found.get("CS5532-ASZR", {}).get("stock_qty") == 700
    assert found.get("LM337IMP", {}).get("stock_qty") == 1_000


# ── 3호 — 병합 셀·전각·비파괴 공백·숨긴 열 ──────────────────────────────────
def _write_messy_cells(tmp_path: Path) -> Path:
    path = tmp_path / "broker-messy-cells.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory"
    sheet.append(["Part Number", "Manufacturer", "Stock", "Lead time", "내부메모"])
    # 병합 셀 — 같은 제조사 두 행을 묶어 첫 행에만 값이 있다.
    sheet.append(["ADN2525ACPZ", "Analog Devices", 300, "Stock", "창고A"])
    sheet.append(["ADN2530YCPZ-500R7", None, 150, "Stock", "창고A"])
    sheet.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    # 전각 숫자·비파괴 공백이 섞인 행.
    sheet.append(["ＴＬＣ２２０２ＣＰＳＲ", "TI", " 450 ", "3day", ""])
    # 숨긴 열이 있어도 값은 읽혀야 한다(사람 눈에만 안 보인다).
    sheet.column_dimensions["E"].hidden = True
    workbook.save(path)
    return path


def test_merged_and_wide_characters_survive(tmp_path: Path):
    result = _run(_write_messy_cells(tmp_path))
    rows = _rows(result)
    assert len(rows) == 3, f"행이 유실됐다: {rows}"

    # ① 병합 셀 아래 행은 제조사가 비어 있다. 채우지 못해도 **행은 남고 재고는 산다**
    #    (제조사 없는 행이 절반인 것이 실측이고, 그래도 저장하는 것이 이 기능의 전제다).
    second = _by_mpn(result).get("ADN2530YCPZ-500R7")
    assert second is not None
    assert second.get("stock_qty") == 150

    # ② 전각 품번은 읽는 단계에서 NFKC 로 반각이 된다(뜻이 보존되는 변환이라 무손실로 본다).
    #    중요한 건 **행이 살아 있고 조회 키가 서는 것** — 전각 그대로 남으면 오히려 안 걸린다.
    wide = _by_mpn(result).get("TLC2202CPSR")
    assert wide is not None, f"전각 품번 행이 사라졌다: {[r.get('part_number') for r in rows]}"
    assert wide.get("stock_qty") == 450, "비파괴 공백이 수량 파싱을 막았다"

    # ③ 숨긴 열도 유실하지 않는다 — 무유실 원칙(미분류 열은 raw_fields 로 간다).
    assert any("창고A" in str(row.get("raw_fields") or {}) for row in rows)

# -*- coding: utf-8 -*-
"""Header-block detection regressions independent of the private BOM corpus."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from bom_extraction_engine.engine import SmartbomConfig, build_smartbom_result
from bom_extraction_engine.normalize import label_form
from bom_extraction_engine.workbook import HeaderNotFound, build_case


def _save_ring_style_header(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RING"
    sheet["B1"] = "PART LIST"

    sheet["B8"] = "NO"
    sheet.merge_cells("B8:B9")
    sheet["C8"] = "DESC/REMARK"
    sheet.merge_cells("C8:F8")
    sheet["G8"] = "MAKER PART NO"
    sheet.merge_cells("G8:G9")
    sheet["H8"] = "Maker"
    sheet.merge_cells("H8:H9")
    sheet["I8"] = "Q'ty"
    sheet.merge_cells("I8:I9")
    sheet["J8"] = "TOP/BOT Reference"
    sheet.merge_cells("J8:L9")
    sheet["M8"] = "REMARK"
    sheet.merge_cells("M8:M9")

    sheet["C9"] = "DESCRIPTION"
    sheet.merge_cells("C9:D9")
    sheet["E9"] = "PACKAGE"
    sheet["F9"] = "VALUE"

    sheet["B10"] = "MAIN CHIP"
    sheet.merge_cells("B10:G10")
    rows = [
        (1, "MCU", "QFN-32", "3.3V", "STM32F030C8T6", "ST", 1, "U1"),
        (2, "LDO", "SOT-23-5", "3.3V", "TPS7A0233PDBVR", "TI", 1, "U2"),
        (3, "LED DRIVER", "QFN-16", "8 channel", "TLC59208", "TI", 1, "U3"),
    ]
    for row_number, values in enumerate(rows, start=11):
        number, description, package, value, mpn, maker, quantity, reference = values
        sheet.cell(row_number, 2, number)
        sheet.cell(row_number, 3, description)
        sheet.cell(row_number, 5, package)
        sheet.cell(row_number, 6, value)
        sheet.cell(row_number, 7, mpn)
        sheet.cell(row_number, 8, maker)
        sheet.cell(row_number, 9, quantity)
        sheet.cell(row_number, 10, reference)
    workbook.save(path)


def test_complementary_second_header_row_is_merged_and_not_emitted(tmp_path: Path):
    source = tmp_path / "two-level-header.xlsx"
    _save_ring_style_header(source)

    case = build_case(source, 0, display_name=source.name, sheet_name="RING")

    assert case["header_rows"] == [7, 8]
    labels_by_column = dict(zip(case["column_indices"], case["header_labels"]))
    assert labels_by_column[2] == "DESC/REMARK/DESCRIPTION"
    assert labels_by_column[4] == "PACKAGE"
    assert labels_by_column[5] == "VALUE"
    assert [row["row_id"] + 1 for row in case["rows"]][:2] == [11, 12]

    result = build_smartbom_result(
        input_path=source,
        original_filename=source.name,
        progress=lambda *_: None,
        config=SmartbomConfig(m2v_path="off"),
    )
    assert result["sheets"][0]["header_rows_1based"] == [8, 9]
    component_rows = {
        row
        for component in result["components"]
        for row in component["source_rows_1based"]
    }
    assert 9 not in component_rows
    assert min(component_rows) == 11


def test_header_like_component_with_numeric_data_is_not_absorbed(tmp_path: Path):
    source = tmp_path / "single-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    sheet.append(["Part Number", "Quantity", "Description"])
    sheet.append(["VALUE", 1, "PACKAGE"])
    sheet.append(["ABC123", 2, "Sensor Module"])
    sheet.append(["DEF456", 3, "Control Module"])
    workbook.save(source)

    case = build_case(source, 0, display_name=source.name, sheet_name="BOM")

    assert case["header_rows"] == [0]
    assert case["rows"][0]["row_id"] == 1


def test_repeated_single_row_header_remains_excluded(tmp_path: Path):
    source = tmp_path / "repeated-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    header = ["Part Number", "Quantity", "Reference"]
    sheet.append(header)
    sheet.append(["ABC123", 1, "U1"])
    sheet.append(["DEF456", 2, "U2 U3"])
    sheet.append(header)
    sheet.append(["GHI789", 1, "U4"])
    workbook.save(source)

    case = build_case(source, 0, display_name=source.name, sheet_name="BOM")

    assert case["header_rows"] == [0, 3]
    assert [row["row_id"] for row in case["rows"]] == [1, 2, 4]


def test_two_cell_complementary_header_is_merged(tmp_path: Path):
    source = tmp_path / "two-cell-continuation.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    sheet.append(["Part Number", "Quantity", "Reference", None, None])
    sheet.append([None, None, None, "Package", "Value"])
    sheet.append(["ABC123", 1, "U1", "QFN-32", "3.3V"])
    sheet.append(["DEF456", 2, "U2 U3", "SOT-23", "5V"])
    workbook.save(source)

    case = build_case(source, 0, display_name=source.name, sheet_name="BOM")

    assert case["header_rows"] == [0, 1]
    labels_by_column = dict(zip(case["column_indices"], case["header_labels"]))
    assert labels_by_column[3] == "Package"
    assert labels_by_column[4] == "Value"
    assert [row["row_id"] for row in case["rows"]] == [2, 3]


@pytest.mark.parametrize(
    "second_anchor",
    [
        ["Part Number", "Quantity", "Reference", None, None, None],
        ["MPN", "Qty", "Ref Des", None, None, None],
    ],
    ids=["identical-anchor", "synonym-anchor"],
)
def test_repeated_two_level_header_blocks_are_excluded(
        tmp_path: Path, second_anchor: list[object]):
    source = tmp_path / "repeated-two-level.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    first_anchor = ["Part Number", "Quantity", "Reference", None, None, None]
    continuation = [None, None, None, "Description", "Package", "Value"]
    sheet.append(first_anchor)
    sheet.append(continuation)
    sheet.append(["ABC123", 1, "U1", "MCU", "QFN-32", "3.3V"])
    sheet.append(["DEF456", 1, "U2", "LDO", "SOT-23", "5V"])
    sheet.append(["SECOND SECTION"])
    sheet.append(second_anchor)
    sheet.append(continuation)
    sheet.append(["GHI789", 2, "U3 U4", "Driver", "QFN-16", "12V"])
    sheet.append(["JKL012", 1, "U5", "Sensor", "LGA-8", "1.8V"])
    workbook.save(source)

    case = build_case(source, 0, display_name=source.name, sheet_name="BOM")

    assert case["header_rows"] == [0, 1, 5, 6]
    assert [row["row_id"] for row in case["rows"]] == [2, 3, 7, 8]
    assert "Part Number" in case["header_labels"]
    assert "Package" in case["header_labels"]


def test_orphan_continuation_does_not_replace_later_header_block(tmp_path: Path):
    source = tmp_path / "orphan-continuation.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    continuation = [None, None, None, "Description", "Package", "Value"]
    sheet.append(["LEGEND"])
    sheet.append(continuation)
    sheet.append([])
    sheet.append(["COMPONENT TABLE"])
    sheet.append([])
    sheet.append(["Part Number", "Quantity", "Reference", None, None, None])
    sheet.append(continuation)
    sheet.append(["ABC123", 1, "U1", "MCU", "QFN-32", "3.3V"])
    sheet.append(["DEF456", 2, "U2 U3", "LDO", "SOT-23", "5V"])
    workbook.save(source)

    case = build_case(source, 0, display_name=source.name, sheet_name="BOM")

    assert case["header_rows"] == [5, 6]
    assert case["header_labels"][0] == "Part Number"
    assert "Package" in case["header_labels"]
    assert [row["row_id"] for row in case["rows"]] == [7, 8]


def test_single_parenthetical_header_continuation_is_merged(tmp_path: Path):
    assert label_form("(Package)") == "package"
    assert label_form("도 면 번 호") == "도면번호"

    source = tmp_path / "single-parenthetical-continuation.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    sheet.append([
        "Index", "Part Name", "Specification", "Type",
        "Vendor", "Qty", "Location",
    ])
    sheet.append([None, None, None, "(Package)"])
    sheet.append([1, "MCU", "STM32F030C8T6", "LQFP-48", "ST", 1, "U1"])
    sheet.append([2, "LDO", "TPS7A0233", "SOT-23", "TI", 1, "U2"])
    workbook.save(source)

    case = build_case(source, 0, display_name=source.name, sheet_name="BOM")

    assert case["header_rows"] == [0, 1]
    labels_by_column = dict(zip(case["column_indices"], case["header_labels"]))
    assert labels_by_column[3] == "Type/(Package)"
    assert [row["row_id"] for row in case["rows"]] == [2, 3]


def test_single_plain_section_label_is_not_a_header_continuation(tmp_path: Path):
    source = tmp_path / "plain-section-label.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    sheet.append(["Part Number", "Quantity", "Reference", "Description"])
    sheet.append([None, None, None, "Description"])
    sheet.append(["ABC123", 1, "U1", "Controller"])
    sheet.append(["DEF456", 1, "U2", "Regulator"])
    workbook.save(source)

    case = build_case(source, 0, display_name=source.name, sheet_name="BOM")

    assert case["header_rows"] == [0]
    assert case["header_labels"] == [
        "Part Number", "Quantity", "Reference", "Description",
    ]


def test_supplier_groups_form_three_row_header_block(tmp_path: Path):
    source = tmp_path / "supplier-groups.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    sheet.append([
        "Item", "Part Number", "Designator", "Quantity", "Supplier",
        None, None, None, None, None, None, None, None, "Footprint",
    ])
    sheet.append([None, None, None, None, "Mouser", None, None,
                  "Digi-Key", None, None, "LCSC.com"])
    sheet.append([
        None, None, None, None, "Mouser No.", "Price/pcs", "Price/set",
        "Digi-Key No.", "Price/pcs", "Price/set", "LCSC Part#",
        "Price/pcs", "Price/set",
    ])
    sheet.append([
        1, "LSM9DS1TR", "U1", 15, "511-LSM9DS1TR", 5.54, 83.1,
        "497-14946-1-ND", 6.104, 91.56, "C12345", 4.5, 67.5,
        "LGA-24",
    ])
    sheet.append([
        2, "LTC4316CMS", "U2", 15, "584-LTC4316CMS", 3.52, 52.8,
        "LTC4316CMS-ND", 3.52, 52.8, "C67890", 2.8, 42.0,
        "MSOP-10",
    ])
    workbook.save(source)

    case = build_case(source, 0, display_name=source.name, sheet_name="BOM")

    assert case["header_rows"] == [0, 1, 2]
    assert [row["row_id"] for row in case["rows"]] == [3, 4]


def test_spaced_korean_labels_are_detected_as_header(tmp_path: Path):
    source = tmp_path / "spaced-korean-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    sheet.append(["B O M"])
    sheet.append([
        "No", "REV", "품 명", "도 면 번 호", "참 고 내 역",
        "Q'TY", "판 매 처", "단 가", "기 타",
    ])
    sheet.append([1, None, "MCU", "U1", "Controller", 1, "Mouser", 100, None])
    sheet.append([2, None, "LDO", "U2", "Regulator", 1, "Digi-Key", 50, None])
    workbook.save(source)

    case = build_case(source, 0, display_name=source.name, sheet_name="BOM")

    assert case["header_rows"] == [1]
    assert [row["row_id"] for row in case["rows"]] == [2, 3]


@pytest.mark.parametrize("include_quantity", [False, True])
def test_generic_duplicate_matrix_is_not_a_bom(
        tmp_path: Path, include_quantity: bool):
    source = tmp_path / "duplicate-check.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "중복검사"
    prefix = ["NO", "Reference"]
    if include_quantity:
        prefix.append("Quantity")
    sheet.append(prefix + [f"열{index}" for index in range(1, 21)] + ["PART NO"])
    for index in range(1, 5):
        row = [index, f"R{index}"]
        if include_quantity:
            row.append(1)
        sheet.append(row + [None] * 20 + [f"RC{index}001"])
    workbook.save(source)

    with pytest.raises(HeaderNotFound):
        build_case(source, 0, display_name=source.name, sheet_name="중복검사")


def test_bom_naming_guide_is_not_a_component_table(tmp_path: Path):
    source = tmp_path / "naming-rule.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Naming Rule"
    sheet.append(["BOM Naming Rule"])
    sheet.append([])
    sheet.append([
        "Comment", "Part / Value", "Designator", "Package",
        "Manufacturer", "Quantity",
    ])
    sheet.append([
        "Comment 할 내용을 적는다", "Part 이름을 표시", "SW : 스위치",
        "패키지 기재", "제조사 기재", 1,
    ])
    sheet.append([
        "설명", "R : R 값 표시", "J : 커넥터", "외형", "공급사", 1,
    ])
    workbook.save(source)

    with pytest.raises(HeaderNotFound):
        build_case(source, 0, display_name=source.name, sheet_name="Naming Rule")
